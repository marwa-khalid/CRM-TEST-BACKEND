from sqlalchemy.orm import Session
from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse,FileResponse
from libdata.models.tables import HireVehicleProvided, HireVehicleStatus
from appflow.models.hire_vehicle_provides import HireVehicleProvidedIn,HireVehicleProvidedUpdateIn
import os
import base64
from openpyxl import Workbook, load_workbook
from io import BytesIO
from sendgrid import SendGridAPIClient
from datetime import datetime
from sendgrid.helpers.mail import (
    Mail, Attachment, FileContent, FileName, FileType, Disposition, ContentId
)
from libdata.models.tables import ClientDetail, Claim,Address,Handler,LocationCondition,VehicleDetail,Borough,Referrer
from libdata.enums import PersonRoleEnum,HistoryLogType
from sqlalchemy.orm import joinedload
from appflow.services.history_activity_service import HistoryActivityService
from appflow.services.graph_email_service import GraphEmailService
from appflow.utils import build_case_reference
from appflow.logger import logger
from sendgrid.helpers.mail import Mail, To, ReplyTo


BASE_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "templates")
FEE_EXEMPTION_FORM_PATH = os.path.join(BASE_TEMPLATE_DIR, "FeeExemptionForm.pdf")
HIRE_TEMPLATE_PATH = os.path.join(BASE_TEMPLATE_DIR, "HireDocumentationAgreement.xlsx")
STORAGE_TEMPLATE_PATH = os.path.join(BASE_TEMPLATE_DIR, "StorageAndRecovery.xlsx")
MITIGATION_TEMPLATE_PATH = os.path.join(BASE_TEMPLATE_DIR, "MitigationQuestionnaire.xlsx")
CHECK_TEMPLATE_PATH = os.path.join(BASE_TEMPLATE_DIR, "VehicleCheckSheet.xlsx")

LOGO_PATH = os.path.join(BASE_TEMPLATE_DIR, "logo.png")
with open(LOGO_PATH, "rb") as logo_file:
    LOGO_ENCODED = base64.b64encode(logo_file.read()).decode()


class HireVehicleProvidedService:

    @staticmethod
    def send_hire_vehicle_email(record, db: Session, current_user: int, tenant_id: int, switch_vehicle: bool = False):
        """Send hire vehicle status email after creation - skip if required fields are null"""

        # Skip email if hire_vehicle_status_id is null
        if not record.hire_vehicle_status_id:
            return

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == record.claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            return
        # if not client:
        #     raise HTTPException(status_code=404, detail=f"No client found for claim ID {record.claim_id}")

        ClientName = client.first_name + " " + client.surname
        if not ClientName:  # If both are None/empty, use a default
            ClientName = "Client"
        our_reference = HireVehicleProvidedService._generate_our_reference(record.claim_id, db)

        # Determine dates for body
        hire_start = record.hire_start_date.strftime("%Y-%m-%d") if record.hire_start_date else "N/A"
        hire_end = record.hire_end_date.strftime("%Y-%m-%d") if record.hire_end_date else "N/A"

        if switch_vehicle:
            # Skip switch vehicle email if required vehicle details are null
            if not all([record.make, record.model, record.hire_vehicle_registration]):
                return

            prev_record = (
                db.query(HireVehicleProvided)
                .filter(
                    HireVehicleProvided.claim_id == record.claim_id,
                    HireVehicleProvided.id < record.id
                )
                .order_by(HireVehicleProvided.id.desc())
                .first()
            )

            if prev_record:
                # Skip if previous record doesn't have required details
                if not all([prev_record.make, prev_record.model, prev_record.hire_vehicle_registration]):
                    return

                prev_start = prev_record.hire_start_date.strftime("%Y-%m-%d") if prev_record.hire_start_date else "N/A"
                prev_end = prev_record.hire_end_date.strftime("%Y-%m-%d") if prev_record.hire_end_date else "N/A"

                html_content = f"""
                <html>
                  <head>
                    <style>
                      body {{ font-family: 'Inter'; margin: 0; padding: 0; }}
                      .header, .footer {{ display: flex; align-items: center; padding: 10px 20px; }}
                      .header {{ border-bottom: 1px solid #ddd; }}
                      .footer {{ border-top: 1px solid #ddd; font-size: 12px; color: #666; justify-content: center; }}
                      .header img {{ width: 40px; height: auto; margin-right: 10px; }}
                      .header p {{ font-size: 20px; font-weight: bold; margin: 0; }}
                    </style>
                  </head>
                  <body>
                    <div class="header">
                      <img src="cid:companylogo" alt="ProClaim" />
                      <p>ProClaim</p>
                    </div>
                    <div style="padding: 20px;">
                      <p>Dear <strong>{ClientName},</strong></p>
                      <p>Our Reference: {our_reference}</p>
                      <p><strong>Previous Vehicle Details:</strong></p>
                      <p>Make: {prev_record.make}</p>
                      <p>Model: {prev_record.model}</p>
                      <p>Registration: {prev_record.hire_vehicle_registration}</p>
                      <p>Hire Start Date: {prev_start}</p>
                      <p>Hire End Date: {prev_end}</p>
                      <br>
                      <p><strong>New Vehicle Details:</strong></p>
                      <p>Make: {record.make}</p>
                      <p>Model: {record.model}</p>
                      <p>Registration: {record.hire_vehicle_registration}</p>
                      <p>Hire Start Date: {record.hire_start_date}</p>
                    </div>
                    <div class="footer">
                              <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span> 
                            </div>
                  </body>
                </html>
                """
                subject = f"Your Hire Vehicle Has Been Switched: {prev_record.hire_vehicle_registration} ➡️ {record.hire_vehicle_registration}"
                HireVehicleProvidedService.send_to_client(record.claim_id, subject, html_content, db, html=True)
                reference = build_case_reference(record.claim_id, db)
                HistoryActivityService.create_activity(
                    db=db,
                    claim_id=record.claim_id,
                    file_name=f"{subject} for claim {reference}",
                    file_path="",
                    file_type=HistoryLogType.VEHICLE_FLEET_ON_SEND,
                    user_id=current_user,
                    tenant_id=tenant_id
                )
            return

        # Not switching → send On Hire / Off Hire
        status = db.query(HireVehicleStatus).filter(HireVehicleStatus.id == record.hire_vehicle_status_id).first()
        if not status:
            return  # Skip email if status not found

        status_label = status.label.lower().strip()

        # Skip email if required vehicle details are null for status-based emails
        if not all([record.make, record.model, record.hire_vehicle_registration]):
            return

        # Compose HTML body
        html_content = f"""
        <html>
          <head>
            <style>
              body {{ font-family: 'Inter'; margin: 0; padding: 0; }}
              .header, .footer {{ display: flex; align-items: center; padding: 10px 20px; }}
              .header {{ border-bottom: 1px solid #ddd; }}
              .footer {{ border-top: 1px solid #ddd; font-size: 12px; color: #666; justify-content: center; }}
              .header img {{ width: 40px; height: auto; margin-right: 10px; }}
              .header p {{ font-size: 20px; font-weight: bold; margin: 0; }}
            </style>
          </head>
          <body>
            <div class="header">
              <img src="cid:companylogo" alt="ProClaim" />
              <p>ProClaim</p>
            </div>
            <div style="padding: 20px;">
              <p><strong>Dear</strong> {ClientName},</p>
              <p><strong>Our Reference:</strong> {our_reference}</p>
              <p><strong>Make:</strong> {record.make}</p>
              <p><strong>Model:</strong> {record.model}</p>
              <p><strong>Registration:</strong> {record.hire_vehicle_registration}</p>
        """

        if "on hire" in status_label:
            subject = f"Your Vehicle is Now On Hire - {record.hire_vehicle_registration}"
            html_content += f"<p><strong>Hire Start Date:</strong> {hire_start}</p>"
        elif "off hire" in status_label:
            subject = f"Your Hire Vehicle is Off Hire - {record.hire_vehicle_registration}"
            html_content += f"<p><strong>Hire Start Date:</strong> {hire_start}</p>"
            html_content += f"<p><strong>Hire End Date:</strong> {hire_end}</p>"
        else:
            return  # Skip email for other statuses

        html_content += """
            </div>
            <div class="footer">
            <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span>
            </div>
          </body>
        </html>
        """

        HireVehicleProvidedService.send_to_client(record.claim_id, subject, html_content, db, html=True)
        HistoryActivityService.create_activity(
            db=db,
            claim_id=record.claim_id,
            file_name=f"{subject} for claim {our_reference}",
            file_path="",
            file_type=HistoryLogType.VEHICLE_FLEET_ON_SEND,
            user_id=current_user,
            tenant_id=tenant_id
        )
    @staticmethod
    def create_hire_vehicle_provided(payload: HireVehicleProvidedIn, db: Session, current_user,tenant_id,switch_vehicle):
        """Create HireVehicleProvided records with validation and automatic email sending"""
        claim_id = payload.claim_id
        section_a_data = payload.section_a.dict(exclude_unset=True)
        created_records = []

        section_b_list = payload.section_b

        for i, item in enumerate(section_b_list):
            HireVehicleProvidedService._validate_hire_vehicle_fields(item, db)

            # Validate multiple Section B hire continuity
            if i > 0:
                prev_item = section_b_list[i - 1]
                if not prev_item.hire_end_date:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Section B record #{i} cannot be created because the previous hire "
                            f"has no 'hire_end_date'. Each new hire requires the previous one to be completed."
                        ),
                    )

            record_data = {
                **(section_a_data if i == 0 else {}),
                **item.dict(),
                "claim_id": claim_id
            }

            record = HireVehicleProvided(**record_data)
            record.created_by = current_user
            record.updated_by = current_user
            db.add(record)
            created_records.append(record)

        db.commit()
        reference = build_case_reference(claim_id,db)
        HistoryActivityService.create_activity(
            db=db,
            claim_id=claim_id,
            file_name=f"The hire vehicle provided detail has been created for claim {reference}",
            file_path="",
            file_type=HistoryLogType.CREATED_HIRE_VEHICLE,
            user_id=current_user,
            tenant_id=tenant_id
        )
        for rec in created_records:
            db.refresh(rec)
        if len(section_b_list) >= 1:
            for rec in created_records:
                # Only send email if the record has complete vehicle details or switch_vehicle is True
                if switch_vehicle or (
                        rec.hire_vehicle_status_id and rec.make and rec.model and rec.hire_vehicle_registration):
                    HireVehicleProvidedService.send_hire_vehicle_email(rec, db, current_user, tenant_id,
                                                                       switch_vehicle=switch_vehicle)

        return created_records

    @staticmethod
    def update_hire_vehicle_provided_by_claim_id(
            claim_id: int,
            payload: HireVehicleProvidedUpdateIn,
            db: Session,
            current_user,
            tenant_id,
            switch_vehicle: bool = False
    ):
        """Update HireVehicleProvided with full create/update/soft-delete logic."""

        existing_records = {
            r.id: r for r in db.query(HireVehicleProvided)
            .filter(
                HireVehicleProvided.claim_id == claim_id,
                HireVehicleProvided.is_active == True,
                HireVehicleProvided.is_deleted == False
            )
            .all()
        }

        incoming_ids = set()
        updated_fields = []
        email_records = []

        section_a = payload.section_a.dict(exclude_unset=True)
        incoming_list = payload.section_b

        # Define field label mapping for human-readable names
        field_label_map = {
            # Section A fields
            "inst_fleet_on_hire": "Inst Fleet On Hire",
            "inst_fleet_off_hire": "Inst Fleet Off Hire",
            "hire_vehicle_check_sheet": "Hire Vehicle Check Sheet",
            "recovery_storage": "Recovery Storage",
            "mitigation_questionnaire": "Mitigation Questionnaire",
            "hire_documentation": "Hire Documentation",
            "fee_exemption_form": "Fee Exemption Form",
            "send_licensing_document_account": "Send Licensing Document Account",
            "request_updated_insurance_schedule": "Request Updated Insurance Schedule",
            "raise_authority_letter": "Raise Authority Letter",

            # Section B fields
            "client_vehicle_category_id": "Client Vehicle Category",
            "actual_vehicle_category_id": "Actual Vehicle Category",
            "cross_hire": "Cross Hire",
            "hire_vehicle_status_id": "Hire Vehicle Status",
            "provider_name": "Provider Name",
            "contact_number": "Contact Number",
            "rate": "Rate",
            "hire_vehicle_registration": "Hire Vehicle Registration",
            "make": "Make",
            "model": "Model",
            "hire_start_date": "Hire Start Date",
            "hire_end_date": "Hire End Date",
            "fuel_type": "Fuel Type",
            "plate_transfer": "Plate Transfer",
        }

        try:
            for item in incoming_list:
                if item.id:
                    incoming_ids.add(item.id)

                    if item.id not in existing_records:
                        continue

                    record = existing_records[item.id]

                    old_status_id = record.hire_vehicle_status_id
                    data = {**section_a, **item.dict(exclude={"id"})}

                    changed_fields = []

                    for key, value in data.items():
                        if not hasattr(record, key):
                            continue

                        old_value = getattr(record, key)
                        if old_value != value:
                            setattr(record, key, value)
                            # Use mapped field name if available, otherwise use the key
                            field_name = field_label_map.get(key, key.replace("_", " ").title())
                            changed_fields.append(field_name)

                    if changed_fields:
                        updated_fields.extend(changed_fields)
                        record.updated_by = current_user

                        if switch_vehicle or old_status_id != record.hire_vehicle_status_id:
                            email_records.append((record, old_status_id))

                else:
                    data = {**section_a, **item.dict(exclude={"id"})}

                    new_record = HireVehicleProvided(
                        **data,
                        claim_id=claim_id,
                        created_by=current_user,
                        updated_by=current_user
                    )
                    db.add(new_record)
                    db.flush()

                    updated_fields.append("New Hire Vehicle Created")
                    email_records.append((new_record, None))

            for r_id, record in existing_records.items():
                if r_id not in incoming_ids:
                    record.is_active = False
                    record.is_deleted = True
                    record.updated_by = current_user
                    updated_fields.append("Hire Vehicle Deleted")

            db.commit()

            for r_id in existing_records:
                db.refresh(existing_records[r_id])
            for record, _ in email_records:
                db.refresh(record)

            # Send emails if needed
            for record, old_status in email_records:
                if switch_vehicle or (
                        record.hire_vehicle_status_id
                        and record.make
                        and record.model
                        and record.hire_vehicle_registration
                ):
                    HireVehicleProvidedService.send_hire_vehicle_email(
                        record, db, current_user, tenant_id, switch_vehicle
                    )

            # Create history activity with formatted field names
            if updated_fields:
                reference = build_case_reference(claim_id, db)
                HistoryActivityService.create_activity(
                    db=db,
                    claim_id=claim_id,
                    file_name=f"The hire vehicle provided details updated for claim {reference}",
                    file_path=", ".join(sorted(set(updated_fields))),  # Remove duplicates and sort
                    file_type=HistoryLogType.UPDATED_HIRE_VEHICLE,
                    user_id=current_user,
                    tenant_id=tenant_id,
                )

            all_records = db.query(HireVehicleProvided).filter(
                HireVehicleProvided.claim_id == claim_id
            ).order_by(HireVehicleProvided.id.asc()).all()

            return all_records

        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Error updating hire vehicle: {str(e)}")
    # @staticmethod
    # def update_hire_vehicle_provided_by_claim_id(
    #         claim_id: int,
    #         payload: HireVehicleProvidedUpdateIn,
    #         db: Session,
    #         current_user,
    #         tenant_id,
    #         switch_vehicle: bool = False
    # ):
    #     """Update HireVehicleProvided with full create/update/soft-delete logic."""
    #
    #     existing_records = {
    #         r.id: r for r in db.query(HireVehicleProvided)
    #         .filter(
    #             HireVehicleProvided.claim_id == claim_id,
    #             HireVehicleProvided.is_active == True,
    #             HireVehicleProvided.is_deleted == False
    #         )
    #         .all()
    #     }
    #
    #     incoming_ids = set()
    #     updated_fields = []
    #     email_records = []
    #
    #     section_a = payload.section_a.dict(exclude_unset=True)
    #     incoming_list = payload.section_b
    #
    #     try:
    #
    #         for item in incoming_list:
    #             if item.id:
    #                 incoming_ids.add(item.id)
    #
    #                 if item.id not in existing_records:
    #                     continue
    #
    #                 record = existing_records[item.id]
    #
    #                 old_status_id = record.hire_vehicle_status_id
    #                 data = {**section_a, **item.dict(exclude={"id"})}
    #
    #                 changed_fields = []
    #
    #                 for key, value in data.items():
    #                     if not hasattr(record, key):
    #                         continue
    #
    #                     old_value = getattr(record, key)
    #                     if old_value != value:
    #                         setattr(record, key, value)
    #                         changed_fields.append(key)
    #
    #                 if changed_fields:
    #                     updated_fields.extend(changed_fields)
    #                     record.updated_by = current_user
    #
    #                     if switch_vehicle or old_status_id != record.hire_vehicle_status_id:
    #                         email_records.append((record, old_status_id))
    #
    #             else:
    #                 data = {**section_a, **item.dict(exclude={"id"})}
    #
    #                 new_record = HireVehicleProvided(
    #                     **data,
    #                     claim_id=claim_id,
    #                     created_by=current_user,
    #                     updated_by=current_user
    #                 )
    #                 db.add(new_record)
    #                 db.flush()
    #
    #                 updated_fields.append("New Hire Vehicle Provided Created")
    #                 email_records.append((new_record, None))
    #
    #         for r_id, record in existing_records.items():
    #             if r_id not in incoming_ids:
    #                 record.is_active = False
    #                 record.is_deleted = True
    #                 record.updated_by = current_user
    #                 updated_fields.append("Hire Vehicle Deleted")
    #
    #         db.commit()
    #
    #         for r_id in existing_records:
    #             db.refresh(existing_records[r_id])
    #         for record, _ in email_records:
    #             db.refresh(record)
    #
    #
    #         for record, old_status in email_records:
    #             if switch_vehicle or (
    #                     record.hire_vehicle_status_id
    #                     and record.make
    #                     and record.model
    #                     and record.hire_vehicle_registration
    #             ):
    #                 HireVehicleProvidedService.send_hire_vehicle_email(
    #                     record, db, current_user, tenant_id, switch_vehicle
    #                 )
    #
    #         if updated_fields:
    #             reference = build_case_reference(claim_id, db)
    #             HistoryActivityService.create_activity(
    #                 db=db,
    #                 claim_id=claim_id,
    #                 file_name=f"The hire vehicle provided details updated for claim {reference}",
    #                 file_path=", ".join(sorted(set(updated_fields))),
    #                 file_type=HistoryLogType.UPDATED_HIRE_VEHICLE,
    #                 user_id=current_user,
    #                 tenant_id=tenant_id,
    #             )
    #
    #         all_records = db.query(HireVehicleProvided).filter(
    #             HireVehicleProvided.claim_id == claim_id
    #         ).order_by(HireVehicleProvided.id.asc()).all()
    #
    #         return all_records
    #
    #     except Exception as e:
    #         db.rollback()
    #         raise HTTPException(500, f"Error updating hire vehicle: {str(e)}")

    @staticmethod
    def _validate_hire_vehicle_fields(data, db: Session):
        """Business rules validation for Hire Vehicle Provided - skip validation for null values"""

        # Skip all validation if hire_vehicle_status_id is null
        if not data.hire_vehicle_status_id:
            return  # Skip validation entirely

        status_obj = db.query(HireVehicleStatus).filter(
            HireVehicleStatus.id == data.hire_vehicle_status_id
        ).first()

        if not status_obj:
            raise HTTPException(
                status_code=404,
                detail=f"Hire Vehicle Status not found (id={data.hire_vehicle_status_id})"
            )

        # Only validate cross_hire requirements if cross_hire is True AND fields are provided
        if data.cross_hire:
            # Check if any required fields are provided but incomplete
            provided_fields = []
            if data.provider_name is not None:
                provided_fields.append("provider_name")
            if data.contact_number is not None:
                provided_fields.append("contact_number")
            if data.rate is not None:
                provided_fields.append("rate")

            # If any cross_hire fields are provided but not all, it's an error
            if provided_fields and len(provided_fields) < 3:
                missing_fields = [f for f in ["provider_name", "contact_number", "rate"] if f not in provided_fields]
                raise HTTPException(
                    status_code=400,
                    detail=f"When providing cross_hire details, all fields {['provider_name', 'contact_number', 'rate']} are required. Missing: {missing_fields}"
                )

        status_name = status_obj.label.lower().strip()

        # Only validate status-specific requirements if the status requires them
        # and the relevant fields are provided (not null)
        if "on hire" in status_name or "vehicle collected" in status_name:
            # Check which fields are provided
            provided_fields = []
            if data.hire_start_date is not None:
                provided_fields.append("hire_start_date")
            if data.make is not None:
                provided_fields.append("make")
            if data.model is not None:
                provided_fields.append("model")
            if data.hire_vehicle_registration is not None:
                provided_fields.append("hire_vehicle_registration")

            # If any status-specific fields are provided but not all, it's an error
            required_fields = ["hire_start_date", "make", "model", "hire_vehicle_registration"]
            if provided_fields and len(provided_fields) < len(required_fields):
                missing_fields = [f for f in required_fields if f not in provided_fields]
                raise HTTPException(
                    status_code=400,
                    detail=f"When providing details for status '{status_name}', all fields {required_fields} are required. Missing: {missing_fields}"
                )

        # Validate hire dates only if both are provided
        if data.hire_start_date and data.hire_end_date:
            if data.hire_end_date < data.hire_start_date:
                raise HTTPException(
                    status_code=400,
                    detail="hire_end_date must be greater than or equal to hire_start_date."
                )
    @staticmethod
    def get_hire_vehicle_provided_by_claim_id(claim_id: int, db: Session, current_user):
        """Get all HireVehicleProvided records for a claim"""
        records = db.query(HireVehicleProvided).filter(
            HireVehicleProvided.claim_id == claim_id,
            HireVehicleProvided.is_active == True
        ).all()

        if not records:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No Hire Vehicle Provided records found for claim_id={claim_id}"
            )
        return records

    @staticmethod
    def deactivate_hire_vehicle_provided(claim_id: int, db: Session, current_user):
        records = db.query(HireVehicleProvided).filter(HireVehicleProvided.claim_id == claim_id).all()
        if not records:
            raise HTTPException(status_code=404, detail="No hire details found for this claim.")

        for record in records:
            record.is_active = False
            record.updated_by = current_user
        db.commit()
        return {"message": f"All hire details for claim_id {claim_id} have been deactivated."}

    @staticmethod
    def get_section_b_vehicle_details(claim_id: int, db: Session):
        records = (
            db.query(
                HireVehicleProvided.make,
                HireVehicleProvided.model,
                HireVehicleProvided.hire_vehicle_registration,
                HireVehicleProvided.hire_start_date,
                HireVehicleProvided.hire_end_date,
            )
            .filter(HireVehicleProvided.claim_id == claim_id,HireVehicleProvided.is_active==True,HireVehicleProvided.is_deleted==False)
            .all()
        )

        if not records:
            raise HTTPException(status_code=404, detail="No hire vehicle records found for this claim.")

        return records

    @staticmethod
    def _generate_our_reference(claim_id: int, db):
        """
        Helper function to generate 'our_reference' string for a given claim.
        Format: {client.surname}-{YYYY}{MM}-{00000ID}
        """
        claim = db.query(Claim).filter(Claim.id == claim_id).first()
        if not claim:
            raise HTTPException(status_code=404, detail=f"No claim found with ID {claim_id}")

        client = db.query(ClientDetail).filter(ClientDetail.claim_id == claim_id, # ClientDetail.role == PersonRoleEnum.CLIENT
                                               ).first()
        # if not client:
        #     raise HTTPException(status_code=404, detail=f"No client found for claim ID {claim_id}")

        year = claim.file_opened_at.strftime("%Y") if claim.file_opened_at else datetime.now().strftime("%Y")
        month = claim.file_opened_at.strftime("%m") if claim.file_opened_at else datetime.now().strftime("%m")
        padded_id = str(claim.id).zfill(5)

        our_reference = f"{client.surname}-{year}{month}-{padded_id}" if client else f"{year}{month}-{padded_id}"
        return our_reference

    @staticmethod
    def send_hire_documentation_agreement_xlsx(claim_id: int, db):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        if not address or not address.email:
            raise HTTPException(status_code=404, detail=f"No email address found for client of claim ID {claim_id}")

        receiver_email = address.email
        address_home = address.address or ""
        Postcode = address.postcode or ""
        Tel_no = address.mobile_tel or ""
        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname
        DOB = client.date_of_birth.strftime("%d-%m-%Y") if client.date_of_birth else "N/A"
        Occupation = client.occupation or "N/A"

        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(HIRE_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {HIRE_TEMPLATE_PATH}")

        wb = load_workbook(HIRE_TEMPLATE_PATH)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = (
                            cell.value
                            .replace("${ClientName}", ClientName)
                            .replace("${DOB}", DOB)
                            .replace("${Occupation}", Occupation)
                            .replace("${address_home}", address_home)
                            .replace("${Postcode}", Postcode)
                            .replace("${Tel_no}", Tel_no)
                        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        xlsx_encoded = base64.b64encode(stream.read()).decode()

        subject = f"Hire Documentation Agreement - Claim No:{our_reference}"
        html_body = f"""
            <html>
              <head>
                <style>
                  body {{
                    font-family: 'Inter';
                    margin: 0;
                    padding: 0;
                  }}
                  .header, .footer {{
                    display: flex;
                    align-items: center;
                    padding: 10px 20px;
                  }}
                  .header {{
                    border-bottom: 1px solid #ddd;
                  }}
                  .footer {{
                    border-top: 1px solid #ddd;
                    font-size: 12px;
                    color: #666;
                    justify-content: center;
                  }}
                  .header img {{
                    width: 40px;
                    height: auto;
                    margin-right: 10px;
                  }}
                  .header p {{
                    font-size: 20px;
                    font-weight: bold;
                    margin: 0;
                  }}
                </style>
              </head>
              <body>
                <div class="header">
                  <img src="cid:companylogo" alt="ProClaim" />
                  <p>ProClaim</p>
                </div>
                <div style="padding: 20px;">
                  <p><strong>Dear {ClientName},</strong></p>
                  <p>Please find attached the <strong>Hire Documentation Agreement</strong>.</p>
                  <p>Regards,<br><strong>Nationwide Assist Team</strong></p>
                </div>
                <div class="footer">
                  <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span> 
                </div>
              </body>
            </html>
            """

        message = Mail(
            from_email="no-replynationwideassist@outlook.com",
            to_emails=receiver_email,
            subject=subject,
            html_content=html_body,
        )

        # Attach filled XLSX
        attachment = Attachment(
            FileContent(xlsx_encoded),
            FileName("HireDocumentationAgreement.xlsx"),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment")
        )

        logo_attachment = Attachment(
            FileContent(LOGO_ENCODED),
            FileName("logo.png"),
            FileType("image/png"),
            Disposition("inline"),
            ContentId("companylogo")
        )

        message.add_attachment(attachment)
        message.add_attachment(logo_attachment)

        api_key = os.getenv("SENDGRID_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="SendGrid API key not found")

        try:
            sg = SendGridAPIClient(api_key)
            response = sg.send(message)
            return {
                "status": "success",
                "message": f"Email with Excel sent to {receiver_email}",
                "sendgrid_status": response.status_code,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def send_storage_recovery_xlsx(claim_id: int, db):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        if not address or not address.email:
            raise HTTPException(status_code=404, detail=f"No email address found for client of claim ID {claim_id}")

        receiver_email = address.email
        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname
        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)


        if not os.path.exists(STORAGE_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {STORAGE_TEMPLATE_PATH}")

        wb = load_workbook(STORAGE_TEMPLATE_PATH)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        xlsx_encoded = base64.b64encode(stream.read()).decode()

        subject = f"Storage and Recovery - Claim No:{our_reference}"
        html_body = f"""
                <html>
                  <head>
                    <style>
                      body {{
                        font-family: 'Inter';
                        margin: 0;
                        padding: 0;
                      }}
                      .header, .footer {{
                        display: flex;
                        align-items: center;
                        padding: 10px 20px;
                      }}
                      .header {{
                        border-bottom: 1px solid #ddd;
                      }}
                      .footer {{
                        border-top: 1px solid #ddd;
                        font-size: 12px;
                        color: #666;
                        justify-content: center;
                      }}
                      .header img {{
                        width: 40px;
                        height: auto;
                        margin-right: 10px;
                      }}
                      .header p {{
                        font-size: 20px;
                        font-weight: bold;
                        margin: 0;
                      }}
                    </style>
                  </head>
                  <body>
                    <div class="header">
                      <img src="cid:companylogo" alt="ProClaim" />
                      <p>ProClaim</p>
                    </div>
                    <div style="padding: 20px;">
                      <p><strong>Dear {ClientName},</strong></p>
                      <p>Please find attached the <strong>Storage And Recovery</strong>.</p>
                      <p>Regards,<br><strong>Nationwide Assist Team</strong></p>
                    </div>
                    <div class="footer">
                      <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span> 
                    </div>
                  </body>
                </html>
                """

        message = Mail(
            from_email="no-replynationwideassist@outlook.com",
            to_emails=receiver_email,
            subject=subject,
            html_content=html_body,
        )

        attachment = Attachment(
            FileContent(xlsx_encoded),
            FileName("StorageAndRecovery.xlsx"),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment")
        )

        logo_attachment = Attachment(
            FileContent(LOGO_ENCODED),
            FileName("logo.png"),
            FileType("image/png"),
            Disposition("inline"),
            ContentId("companylogo")
        )

        message.add_attachment(attachment)
        message.add_attachment(logo_attachment)

        api_key = os.getenv("SENDGRID_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="SendGrid API key not found")

        try:
            sg = SendGridAPIClient(api_key)
            response = sg.send(message)
            return {
                "status": "success",
                "message": f"Email with Excel sent to {receiver_email}",
                "sendgrid_status": response.status_code,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def send_mitigation_questionnaire_xlsx(claim_id: int, db):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        if not address or not address.email:
            raise HTTPException(status_code=404, detail=f"No email address found for client of claim ID {claim_id}")

        handler = (
            db.query(Handler).filter(Handler.id == Claim.handler_id).first()
        )

        location = (
            db.query(LocationCondition).filter(LocationCondition.claim_id == claim_id).first()
        )

        receiver_email = address.email
        address_home = address.address or ""
        Tel_no = address.mobile_tel or ""
        Postcode = address.postcode or ""
        location_str = location.date_time.strftime("%d-%m-%Y %H:%M") if location and location.date_time else "N/A"
        handler = handler.label
        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname
        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(MITIGATION_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {MITIGATION_TEMPLATE_PATH}")

        wb = load_workbook(MITIGATION_TEMPLATE_PATH)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = (
                            cell.value
                            .replace("${ClientName}", ClientName)
                            .replace("${handler}", handler)
                            .replace("${address_home}", address_home)
                            .replace("${Postcode}", Postcode)
                            .replace("${Tel_no}", Tel_no)
                            .replace("${location}", location_str)
                        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        xlsx_encoded = base64.b64encode(stream.read()).decode()

        subject = f"Mitigation Questionnaire - Claim No:{our_reference}"
        html_body = f"""
                    <html>
                      <head>
                        <style>
                          body {{
                            font-family: 'Inter';
                            margin: 0;
                            padding: 0;
                          }}
                          .header, .footer {{
                            display: flex;
                            align-items: center;
                            padding: 10px 20px;
                          }}
                          .header {{
                            border-bottom: 1px solid #ddd;
                          }}
                          .footer {{
                            border-top: 1px solid #ddd;
                            font-size: 12px;
                            color: #666;
                            justify-content: center;
                          }}
                          .header img {{
                            width: 40px;
                            height: auto;
                            margin-right: 10px;
                          }}
                          .header p {{
                            font-size: 20px;
                            font-weight: bold;
                            margin: 0;
                          }}
                        </style>
                      </head>
                      <body>
                        <div class="header">
                          <img src="cid:companylogo" alt="ProClaim" />
                          <p>ProClaim</p>
                        </div>
                        <div style="padding: 20px;">
                          <p><strong>Dear {ClientName},</strong></p>
                          <p>Please find attached the <strong>Mitigation Questionnaire</strong>.</p>
                          <p>Regards,<br><strong>Nationwide Assist Team</strong></p>
                        </div>
                        <div class="footer">
                          <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span> 
                        </div>
                      </body>
                    </html>
                    """

        message = Mail(
            from_email="no-replynationwideassist@outlook.com",
            to_emails=receiver_email,
            subject=subject,
            html_content=html_body,
        )

        attachment = Attachment(
            FileContent(xlsx_encoded),
            FileName("MitigationQuestionnaire.xlsx"),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment")
        )

        logo_attachment = Attachment(
            FileContent(LOGO_ENCODED),
            FileName("logo.png"),
            FileType("image/png"),
            Disposition("inline"),
            ContentId("companylogo")
        )

        message.add_attachment(attachment)
        message.add_attachment(logo_attachment)

        api_key = os.getenv("SENDGRID_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="SendGrid API key not found")

        try:
            sg = SendGridAPIClient(api_key)
            response = sg.send(message)
            return {
                "status": "success",
                "message": f"Email with Excel sent to {receiver_email}",
                "sendgrid_status": response.status_code,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def send_vehicle_check_sheet_xlsx(claim_id: int, db):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        if not address or not address.email:
            raise HTTPException(status_code=404, detail=f"No email address found for client of claim ID {claim_id}")


        receiver_email = address.email
        address_home = address.address or ""
        Tel_no = address.mobile_tel or ""
        Postcode = address.postcode or ""
        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname
        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(CHECK_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {CHECK_TEMPLATE_PATH}")

        wb = load_workbook(CHECK_TEMPLATE_PATH)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = (
                            cell.value
                            .replace("${ClientName}", ClientName)
                            .replace("${address_home}", address_home)
                            .replace("${Postcode}", Postcode)
                            .replace("${Tel_no}", Tel_no)
                        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        xlsx_encoded = base64.b64encode(stream.read()).decode()

        subject = f"Hire Vehicle Check Sheet - Claim No:{our_reference}"
        html_body = f"""
                        <html>
                          <head>
                            <style>
                              body {{
                                font-family: 'Inter';
                                margin: 0;
                                padding: 0;
                              }}
                              .header, .footer {{
                                display: flex;
                                align-items: center;
                                padding: 10px 20px;
                              }}
                              .header {{
                                border-bottom: 1px solid #ddd;
                              }}
                              .footer {{
                                border-top: 1px solid #ddd;
                                font-size: 12px;
                                color: #666;
                                justify-content: center;
                              }}
                              .header img {{
                                width: 40px;
                                height: auto;
                                margin-right: 10px;
                              }}
                              .header p {{
                                font-size: 20px;
                                font-weight: bold;
                                margin: 0;
                              }}
                            </style>
                          </head>
                          <body>
                            <div class="header">
                              <img src="cid:companylogo" alt="ProClaim" />
                              <p>ProClaim</p>
                            </div>
                            <div style="padding: 20px;">
                              <p><strong>Dear {ClientName},</strong></p>
                              <p>Please find attached the <strong>Hire Vehicle Check Sheet</strong>.</p>
                              <p>Regards,<br><strong>Nationwide Assist Team</strong></p>
                            </div>
                            <div class="footer">
                              <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span> 
                            </div>
                          </body>
                        </html>
                        """

        message = Mail(
            from_email="no-replynationwideassist@outlook.com",
            to_emails=receiver_email,
            subject=subject,
            html_content=html_body,
        )

        attachment = Attachment(
            FileContent(xlsx_encoded),
            FileName("HireVehicleCheckSheet.xlsx"),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment")
        )

        logo_attachment = Attachment(
            FileContent(LOGO_ENCODED),
            FileName("logo.png"),
            FileType("image/png"),
            Disposition("inline"),
            ContentId("companylogo")
        )

        message.add_attachment(attachment)
        message.add_attachment(logo_attachment)

        api_key = os.getenv("SENDGRID_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="SendGrid API key not found")

        try:
            sg = SendGridAPIClient(api_key)
            response = sg.send(message)
            return {
                "status": "success",
                "message": f"Email with Excel sent to {receiver_email}",
                "sendgrid_status": response.status_code,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @staticmethod
    def download_fee_exemption_form(claim_id: int, db,current_user: int,tenant: int):

        # # Optional: validate claim exists
        # client = (
        #     db.query(ClientDetail)
        #     .filter(ClientDetail.claim_id == claim_id)
        #     .first()
        # )
        # if not client:
        #     raise HTTPException(status_code=404, detail=f"No client found for claim ID {claim_id}")

        pdf_path = FEE_EXEMPTION_FORM_PATH

        if not os.path.exists(pdf_path):
            raise HTTPException(status_code=404, detail=f"Fee Exemption Form PDF not found")

        reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)
        HistoryActivityService.create_activity(
            db=db,
            claim_id=claim_id,
            file_name=f"The fee exemption pdf downloaded for claim {reference}",
            file_path="",
            file_type=HistoryLogType.DOWNLOAD_EXEMPTION_PDF,
            user_id=current_user,
            tenant_id=tenant
        )
        return FileResponse(
            path=pdf_path,
            filename=f"Fee_Exemption_Form_{reference}.pdf",
            media_type="application/pdf"
        )

    @staticmethod
    def send_fee_exemption_email(claim_id: int, db):

        client = (
            db.query(ClientDetail)
            .filter(ClientDetail.claim_id == claim_id)
            .first()
        )

        if not client:
            raise HTTPException(status_code=404, detail=f"No client found for claim ID {claim_id}")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        if not address or not address.email:
            raise HTTPException(status_code=404, detail=f"No email found for client of claim ID {claim_id}")

        receiver_email = address.email
        client_name = f"{client.first_name} {client.surname}".strip() or "Client"
        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        pdf_path = FEE_EXEMPTION_FORM_PATH
        logo_path = LOGO_PATH

        if not os.path.exists(pdf_path):
            raise HTTPException(status_code=404, detail=f"Fee Exemption Form PDF not found at {pdf_path}")

        if not os.path.exists(logo_path):
            raise HTTPException(status_code=404, detail=f"Logo not found at {logo_path}")

        with open(pdf_path, "rb") as f:
            pdf_encoded = base64.b64encode(f.read()).decode()

        with open(logo_path, "rb") as f:
            logo_encoded = base64.b64encode(f.read()).decode()

        html_content = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: 'Inter'; margin: 0; padding: 0; background-color: #fff; }}
                    .header, .footer {{ display: flex; align-items: center; padding: 10px 20px; }}
                    .header {{ border-bottom: 1px solid #ddd; }}
                    .footer {{ border-top: 1px solid #ddd; font-size: 12px; color: #666; justify-content: center; }}
                    .header img {{ width: 40px; height: auto; margin-right: 10px; }}
                    .header p {{ font-size: 20px; font-weight: bold; margin: 0; }}
                    .content {{ padding: 25px; font-size: 15px; color: #333; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <img src="cid:companylogo" alt="ProClaim" />
                    <p>ProClaim</p>
                </div>
                <div class="content">
                    <p><strong>Dear {client_name},</strong></p>
                    <p>Please find attached the <strong>Fee Exemption Form</strong>.</p>
                    <p>Regards,<br><strong>Nationwide Assist Team</strong></p>
                </div>
                <div class="footer">
                              <p>This email was sent to claim@nationwideassist.co.uk. If you'd rather not receive this kind of email, you can unsubscribe or manage your email preferences.<br> <img src="cid:companylogo" alt="ProClaim" width="25" style="vertical-align: middle; margin-right: 6px;" /> <span style="font-size:13px; font-weight:bold; vertical-align: middle;">ProClaim</span> 
                            </div>
            </body>
            </html>
            """

        message = Mail(
            from_email="no-replynationwideassist@outlook.com",
            to_emails=receiver_email,
            subject=f"Fee Exemption Form - Claim No:{our_reference}",
            html_content=html_content,
        )

        message.add_attachment(
            Attachment(
                FileContent(pdf_encoded),
                FileName("FeeExemptionForm.pdf"),
                FileType("application/pdf"),
                Disposition("attachment"),
            )
        )

        logo_attachment = Attachment(
            FileContent(logo_encoded),
            FileName("logo.png"),
            FileType("image/png"),
            Disposition("inline"),
            ContentId("companylogo"),
        )
        message.add_attachment(logo_attachment)

        api_key = os.getenv("SENDGRID_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="SendGrid API key not found in environment")

        sg = SendGridAPIClient(api_key)
        response = sg.send(message)

        return {
            "status": "success",
            "message": f"Email sent successfully to {receiver_email}",
            "client_name": client_name,
            "claim_id": claim_id,
            "sendgrid_status": response.status_code,
            "timestamp": datetime.now().isoformat(),
        }

    @staticmethod
    def get_instruct_fleet_summary(
            claim_id: int,
            db: Session,
            on_hire: bool = False,
            off_hire: bool = False,
            send_fee_exemption_form: bool = False,
            hire_documentation_agreement: bool = False,
            storage_recovery: bool = False,
            mitigation_questionnaire: bool = False,
            vehicle_check_sheet: bool = False
    ):
        claim = (
            db.query(Claim)
            .options(joinedload(Claim.claim_type))
            .filter(Claim.id == claim_id)
            .first()
        )
        if not claim:
            raise HTTPException(status_code=404, detail="Claim not found")

        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        referrer = db.query(Referrer).filter(Referrer.claim_id == claim_id).first()

        client = (
            db.query(ClientDetail)
            .filter(ClientDetail.claim_id == claim_id, ClientDetail.role == "CLIENT")
            .first()
        )

        address = db.query(Address).filter(Address.id == client.address_id).first() if client else None


        def base_response(subject: str):
            to_recipients = "ayesha.rana@nationwideassist.co.uk;ayesha.rana@nationwideassist.co.uk"
            return {
                "claim_id": claim.id,
                "to": to_recipients,
                "Subject": subject,
                "Reference": our_reference,
                "Referrer": (referrer.company_name or referrer.contact_name) if referrer else None,
                "client_name": f"{client.first_name} {client.surname}" if client else None,
                "mobile_tel": address.mobile_tel if address else None,
            }

        if off_hire:
            return base_response(f"New Instruction to Fleet to Off Hire Vehicle (CIL) - {our_reference}")
        if send_fee_exemption_form:
            return base_response(f"Fee Exemption Form - {our_reference}")
        if hire_documentation_agreement:
            return base_response(f"Hire Documentation Agreement - {our_reference}")
        if storage_recovery:
            return base_response(f"Storage and Recovery - {our_reference}")
        if mitigation_questionnaire:
            return base_response(f"Mitigation Questionnaire - {our_reference}")
        if vehicle_check_sheet:
            return base_response(f"Hire Vehicle Check Sheet - {our_reference}")

        # ✅ on_hire mode (full details)
        if on_hire:
            vehicle = (
                db.query(VehicleDetail)
                .options(
                    joinedload(VehicleDetail.fuel_type),
                    joinedload(VehicleDetail.transmission),
                    joinedload(VehicleDetail.borough).joinedload(Borough.taxi_type),
                )
                .filter(VehicleDetail.claim_id == claim_id)
                .first()
            )

            return {
                **base_response(f"New Instruction to Fleet to Arrange New Hire - {our_reference}"),
                "driver_base": client.driver_base if client else None,
                "make": vehicle.make if vehicle else None,
                "model": vehicle.model if vehicle else None,
                "registration": vehicle.registration if vehicle else None,
                "Auto": vehicle.transmission.label if vehicle and vehicle.transmission else None,
                "fuel_type": vehicle.fuel_type.label if vehicle and vehicle.fuel_type else None,
                "engine_size": vehicle.engine_size if vehicle else None,
                "number_of_seat": vehicle.number_of_seat if vehicle else None,
                "body_type": vehicle.body_type if vehicle else None,
                "borough_name": vehicle.borough.borough_name if vehicle and vehicle.borough else None,
                "taxi_type": vehicle.borough.taxi_type.label if vehicle and vehicle.borough and vehicle.borough.taxi_type else None,
            }

        return base_response("Instruction Summary")

    @staticmethod
    def download_hire_documentation_agreement_xlsx(claim_id: int, db,current_user:int,tenant_id:int):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        address_home = (address.address or "") if address else ""
        Postcode = (address.postcode or "") if address else ""
        Tel_no = (address.mobile_tel or "") if address else ""
        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname
        DOB = client.date_of_birth.strftime("%d-%m-%Y") if client.date_of_birth else "N/A"
        Occupation = client.occupation or "N/A"
        last_hire = (db.query(HireVehicleProvided).filter(HireVehicleProvided.claim_id == claim_id).order_by(HireVehicleProvided.id.desc()).first())
        actual_vehicle_category = (last_hire.actual_vehicle_category.label if last_hire and last_hire.actual_vehicle_category else "")
        hire_start_date = (last_hire.hire_start_date.strftime("%d-%m-%Y") if last_hire and last_hire.hire_start_date else "")
        hire_end_date = (last_hire.hire_end_date.strftime("%d-%m-%Y") if last_hire and last_hire.hire_end_date else "")
        make = (last_hire.make if last_hire and last_hire.make else "")
        model = (last_hire.model if last_hire and last_hire.model else "")
        registration = (last_hire.hire_vehicle_registration if last_hire and last_hire.hire_vehicle_registration else "")

        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(HIRE_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {HIRE_TEMPLATE_PATH}")

        wb = load_workbook(HIRE_TEMPLATE_PATH)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = (
                            cell.value
                            .replace("${ClientName}", ClientName)
                            .replace("${DOB}", DOB)
                            .replace("${Occupation}", Occupation)
                            .replace("${address_home}", address_home)
                            .replace("${Postcode}", Postcode)
                            .replace("${Tel_no}", str(Tel_no) if Tel_no else "")
                            .replace("${Reference}", our_reference)
                            .replace("${ActualCategory}", actual_vehicle_category)
                            .replace("${HireStart}", hire_start_date)
                            .replace("${HireEnd}", hire_end_date)
                            .replace("${make}", make)
                            .replace("${model}", model)
                            .replace("${registration}",registration)
                        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"HireDocumentationAgreement_{our_reference}.xlsx"
        HistoryActivityService.create_activity(
            db=db,
            claim_id=claim_id,
            file_name=f"The hire documentation agreement xlsx downloaded for claim {our_reference}",
            file_path="",
            file_type=HistoryLogType.DOWNLOAD_HIRE_XLSX,
            user_id=current_user,
            tenant_id=tenant_id
        )

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    @staticmethod
    def download_storage_recovery_xlsx(claim_id: int, db,current_user:int,tenant_id:int):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname

        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(STORAGE_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {STORAGE_TEMPLATE_PATH}")

        wb = load_workbook(STORAGE_TEMPLATE_PATH)

        # TODO: Fill placeholders if needed
        # currently, template untouched

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"StorageAndRecovery_{our_reference}.xlsx"
        HistoryActivityService.create_activity(
            db=db,
            claim_id=claim_id,
            file_name=f"The storage and recovery xlsx downloaded for claim {our_reference}",
            file_path="",
            file_type=HistoryLogType.DOWNLOAD_RECOVERY_XLSX,
            user_id=current_user,
            tenant_id=tenant_id
        )

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    @staticmethod
    def download_mitigation_questionnaire_xlsx(claim_id: int, db,current_user:int,tenant_id:int):
        claim = db.query(Claim).filter(Claim.id == claim_id).first()
        if not claim:
            raise HTTPException(status_code=404, detail="Claim not found")

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        handler_obj = (
            db.query(Handler).filter(Handler.id == claim.handler_id).first()
        )

        location = db.query(LocationCondition).filter(
            LocationCondition.claim_id == claim_id
        ).first()

        ClientName = f"{client.first_name} {client.surname}" if client.first_name else client.surname
        address_home = (address.address or "") if address else ""
        Tel_no = (address.mobile_tel or "") if address else ""
        Postcode = (address.postcode or "") if address else ""
        handler = handler_obj.label if handler_obj else ""
        location_str = (
            location.date_time.strftime("%d-%m-%Y %H:%M")
            if location and location.date_time
            else "N/A"
        )

        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(MITIGATION_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {MITIGATION_TEMPLATE_PATH}")

        wb = load_workbook(MITIGATION_TEMPLATE_PATH)

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = (
                            cell.value
                            .replace("${ClientName}", ClientName)
                            .replace("${handler}", handler)
                            .replace("${address_home}", address_home)
                            .replace("${Postcode}", Postcode)
                            .replace("${Tel_no}", str(Tel_no) if Tel_no else "")
                            .replace("${location}", location_str)
                            .replace("${reference_claim}", our_reference)
                        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"MitigationQuestionnaire_{our_reference}.xlsx"
        HistoryActivityService.create_activity(
            db=db,
            claim_id=claim_id,
            file_name=f"The mitigation questionnaire xlsx downloaded for claim {our_reference}",
            file_path="",
            file_type=HistoryLogType.DOWNLOAD_MITIGATION_XLSX,
            user_id=current_user,
            tenant_id=tenant_id
        )

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    @staticmethod
    def download_vehicle_check_sheet_xlsx(claim_id: int, db,current_user:int,tenant_id:int):

        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
            # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = (
            db.query(Address)
            .filter(Address.id == client.address_id)
            .first()
        )

        if not address:
            raise HTTPException(status_code=404, detail="Address not found")

        address_home = address.address or ""
        Tel_no = address.mobile_tel or ""
        Postcode = address.postcode or ""
        ClientName = f"{client.first_name or ''} {client.surname or ''}".strip()
        our_reference = HireVehicleProvidedService._generate_our_reference(claim_id, db)

        if not os.path.exists(CHECK_TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template XLSX not found at {CHECK_TEMPLATE_PATH}")

        wb = load_workbook(CHECK_TEMPLATE_PATH)

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = (
                            cell.value
                            .replace("${ClientName}", ClientName)
                            .replace("${address_home}", address_home)
                            .replace("${Postcode}", Postcode)
                            .replace("${Tel_no}", Tel_no)
                            .replace("${Refernece}",our_reference)
                        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"HireVehicleCheckSheet_{our_reference}.xlsx"
        HistoryActivityService.create_activity(
            db=db,
            claim_id=claim_id,
            file_name=f"The hire vehicle check sheet xlsx downloaded for claim {our_reference}",
            file_path="",
            file_type=HistoryLogType.DOWNLOAD_VEHICLE_XLSX,
            user_id=current_user,
            tenant_id=tenant_id
        )

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    @staticmethod
    def send_to_client(claim_id: int, subject: str, body: str, db: Session, html: bool = False):
        """
        Sends an email to the client for a given claim using SendGrid.
        :param claim_id: Claim ID
        :param subject: Email subject
        :param body: Email body
        :param db: DB session
        :param html: If True, send as HTML, else plain text
        """
        # Fetch client email
        client = db.query(ClientDetail).filter(
            ClientDetail.claim_id == claim_id,
             # ClientDetail.role == PersonRoleEnum.CLIENT
        ).first()

        if not client:
            return
            # raise HTTPException(status_code=404, detail="Client not found for this claim ID")

        address = db.query(Address).filter(Address.id == client.address_id).first()
        if not address or not address.email:
            return  # Skip email if no email address
        # if not address or not address.email:
        #     raise HTTPException(status_code=404, detail=f"No email address found for client of claim ID {claim_id}")

        receiver_email = address.email
        sender_email = "no-replynationwideassist@outlook.com"  # your sender email

        # Compose message
        message = Mail(
            from_email=sender_email,
            to_emails=receiver_email,
            subject=subject,
            html_content=body if html else None,
            plain_text_content=body if not html else None
        )

        # Optionally, add inline logo (if you want)
        # if os.path.exists("logo.png"):
        #     with open("logo.png", "rb") as f:
        #         logo_encoded = base64.b64encode(f.read()).decode()
        logo_attachment = Attachment(
                FileContent(LOGO_ENCODED),
                FileName("logo.png"),
                FileType("image/png"),
                Disposition("inline"),
                ContentId("companylogo")
            )
        message.add_attachment(logo_attachment)

        api_key = os.getenv("SENDGRID_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="SendGrid API key not found")

        try:
            sg = SendGridAPIClient(api_key)
            response = sg.send(message)
            print(f"Email sent to {receiver_email} for claim {claim_id}, status: {response.status_code}")
            return {
                "status": "success",
                "message": f"Email sent to {receiver_email}",
                "sendgrid_status": response.status_code,
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"SendGrid send error: {str(e)}")

    @staticmethod
    def _split_email_recipients(recipients: str):
        raw_recipients = str(recipients or "").replace(",", ";").split(";")
        recipient_list = []
        seen = set()
        for email in raw_recipients:
            clean_email = email.strip()
            key = clean_email.lower()
            if clean_email and key not in seen:
                recipient_list.append(clean_email)
                seen.add(key)
        return recipient_list

    @classmethod
    def _send_email(
        cls,
        recipients: str,
        subject: str,
        html_content: str,
        copy_recipients: str = "",
    ):
        configured_copies = ";".join(
            [
                os.getenv("HIRE_INSTRUCTION_COPY_EMAIL", ""),
                os.getenv("MS_GRAPH_MAILBOX", ""),
                os.getenv("OUTLOOK_MAILBOX", ""),
            ]
        )
        recipient_list = cls._split_email_recipients(
            ";".join(
                [
                    str(recipients or ""),
                    str(copy_recipients or ""),
                    configured_copies,
                ]
            )
        )
        if not recipient_list:
            logger.warning("email skipped: no recipients provided")
            return None

        from_email = os.getenv("SENDGRID_SENDER", "no-replynationwideassist@outlook.com")
        reply_to = os.getenv("SENDGRID_REPLY_TO", from_email)

        # Prefer Microsoft Graph: it sends from a real Outlook mailbox, so mail
        # is actually delivered. SendGrid from the unverified yopmail.com sender
        # gets silently dropped by strict providers (Outlook). Fall back to
        # SendGrid only if Graph isn't configured or the Graph send fails.
        if GraphEmailService.is_configured():
            result = GraphEmailService.send_mail(
                recipient_list, subject, html_content, reply_to=reply_to
            )
            if result is not None:
                return result
            logger.warning("Graph send failed; falling back to SendGrid")

        message = Mail(
            from_email=from_email,
            to_emails=recipient_list,  # SendGrid handles a list of strings here
            subject=subject,
            html_content=html_content
        )

        message.reply_to = ReplyTo(reply_to, "No-Reply")

        # Inline logo so cid:companylogo renders on the SendGrid fallback path too.
        if LOGO_ENCODED and "cid:companylogo" in (html_content or ""):
            message.add_attachment(Attachment(
                FileContent(LOGO_ENCODED), FileName("logo.png"),
                FileType("image/png"), Disposition("inline"), ContentId("companylogo"),
            ))

        try:
            api_key = os.getenv("SENDGRID_API_KEY")
            sg = SendGridAPIClient(api_key)
            response = sg.send(message)
            logger.info(
                "Hire instruction email accepted by SendGrid: "
                f"status={response.status_code}, recipients={recipient_list}"
            )
            return response
        except Exception as e:
            logger.warning(f"SendGrid Error: {e}")
            return None  # Return None so the router can throw an HTTPException
        
    @classmethod
    def process_on_hire_instruction(cls, data: dict):
        subject = data.get('Subject', "New Instruction to Fleet to Arrange New Hire")
        start_date = datetime.now().strftime("%d/%m/%Y")
        
        # Building the dynamic HTML sections
        html_content = f"""
        <div style="font-family: Arial, sans-serif; background-color: #ffffff; padding: 20px; color: #334155;">
            <div style="text-align: center; margin-bottom: 30px;">
                <img src="cid:companylogo" alt="Logo" style="width: 150px;">
            </div>
            
            <div style="max-width: 420px; margin: 0 auto 20px auto; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px;">
                {cls._gen_row("Brand", "RTA - Nationwide Assist")}
                {cls._divider()}
                {cls._gen_row("Reference", data.get('Reference'))}
                {cls._divider()}
                {cls._gen_row("Referrer", data.get('Referrer'))}
                {cls._divider()}
                {cls._gen_row("Client", data.get('client_name'))}
                {cls._divider()}
                {cls._gen_row("Cl Mobile No", data.get('mobile_tel'))}
                <p style="text-align:center; font-size:12px; font-weight:600; margin-top:10px;">Does Hirer Require Vehicle Documents: Yes</p>
            </div>

            <div style="max-width: 420px; margin: 0 auto 20px auto; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px;">
                <p style="font-weight:700; font-size:13px; margin-bottom:10px;">Client's Vehicle Details</p>
                {cls._gen_row("Reg", data.get('registration'))}
                {cls._divider()}
                {cls._gen_row("Make/Model", f"{data.get('make')} / {data.get('model')}")}
                {cls._divider()}
                {cls._gen_row("Body Type", data.get('body_type'))}
                {cls._divider()}
                {cls._gen_row("Auto", data.get('Auto'))}
                {cls._divider()}
                {cls._gen_row("Engine Size", data.get('engine_size'))}
                {cls._divider()}
                {cls._gen_row("Fuel Type", data.get('fuel_type'))}
                {cls._divider()}
                {cls._gen_row("No of Seats", data.get('number_of_seat'))}
            </div>
        """

        # Add Taxi Section if relevant
        if data.get('borough_name') or data.get('taxi_type') or data.get('driver_base'):
            html_content += f"""
            <div style="max-width: 420px; margin: 0 auto 20px auto; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px;">
                <p style="font-weight:700; font-size:13px; margin-bottom:10px;">If Taxi Vehicle</p>
                {cls._gen_row("Borough", data.get('borough_name'))}
                {cls._divider()}
                {cls._gen_row("Type of Plate", data.get('taxi_type'))}
                {cls._divider()}
                {cls._gen_row("Driver Base", data.get('driver_base'))}
            </div>
            """

        # Add Footer Message
        html_content += f"""
            <div style="text-align: center; margin: 30px auto; max-width: 350px; font-size: 14px; line-height: 1.6;">
                <p>Hi,<br><br>
                Please contact the Client to arrange a hire vehicle.<br><br>
                Hire needs to start on <strong>{start_date}</strong>.<br><br>
                We need to provide hire vehicle category <strong>{data.get('vehicleCategory', 'T12<3YRS')}</strong>.</p>
            </div>
            <div style="text-align: center; font-size: 12px; border-top: 1px solid #eee; padding-top: 20px;">
                <p style="font-weight: 600;">Kind regards,<br>Nationwide Assist IT / Systems Team</p>
            </div>
        </div>
        """
        return cls._send_email(
            data.get("to", ""),
            subject,
            html_content,
            data.get("activeUserEmail", ""),
        )

    @staticmethod
    def _gen_row(label, val):
        return f'<div style="display:flex; justify-content:space-between; padding:5px 0;"><span style="font-size:11px; color:#666; width:140px;">{label}</span><span style="font-size:11px; font-weight:700; text-align:left; flex:1;">{val or "N/A"}</span></div>'

    @staticmethod
    def _divider():
        return '<div style="height:1px; background-color:#f1f1f1; margin:2px 0;"></div>'
        
    @classmethod
    def process_off_hire_instruction(cls, data: dict):
        """
        Processes data from frontend and triggers the off-hire email.
        """
        subject = "New Instruction to Fleet to Off Hire Vehicle (CIL)"
        
        # 2. Format Date (DD/MM/YYYY)
        off_hire_date = datetime.now().strftime("%d/%m/%Y")
        # 3. Construct HTML Template (Matching your React Design)
        # Note: We use inline styles because email clients ignore external CSS/Tailwind
        html_content = f"""
        <div style="font-family: 'Arial', sans-serif; background-color: #ffffff; padding: 40px; color: #334155;">
        <div style="text-align: center; margin-bottom: 20px;">
            <img src="cid:companylogo" alt="Logo" style="width: 150px;">
        </div>
        
        <div style="max-width: 384px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; background-color: #ffffff;">
            {cls._generate_data_row("Reference", data.get('Reference'))}
            {cls._generate_divider()}
            {cls._generate_data_row("Referrer", data.get('Referrer'))}
            {cls._generate_divider()}
            {cls._generate_data_row("Client", data.get('client_name'))}
            {cls._generate_divider()}
            {cls._generate_data_row("Hire Vehicle", data.get('registration'))}
            {cls._generate_divider()}
            {cls._generate_data_row("Cl Mobile No", data.get('mobile_tel'))}
        </div>

        <div style="text-align: center; margin: 40px auto; max-width: 320px; color: #334155;">
            <p style="font-size: 14px; font-weight: 400; line-height: 1.5;">
                Hi,<br><br>
                Please contact the Client to arrange the off hire of this vehicle for 
                <span style="font-weight: 600;">{off_hire_date}</span>.
            </p>
        </div>

        <div style="max-width: 580px; height: 1px; background-color: #e2e8f0; margin: 20px auto;"></div>

        <div style="text-align: center; color: #334155;">
            <p style="font-size: 12px; font-weight: 600; margin-bottom: 4px;">Kind regards,</p>
            <p style="font-size: 14px; font-weight: 600;">Nationwide Assist<br>IT / Systems Team</p>
        </div>
    </div>
    """
        # 4. Call your existing static method
        return cls._send_email(
            data.get("to", ""),
            subject,
            html_content,
            data.get("activeUserEmail", ""),
        )

    @staticmethod
    def _generate_data_row(label, value):
        # Label: Normal weight (400), Slate-500/700, XS size
        # Value: SemiBold (600), Slate-700, XS size
        return f"""
        <div style="display: flex; align-items: center; padding: 8px 0;">
            <div style="width: 130px; color: #334155; font-size: 12px; font-weight: 400; font-family: 'Arial', sans-serif;">{label}</div>
            <div style="color: #334155; font-size: 12px; font-weight: 600; font-family: 'Arial', sans-serif;">{value or 'N/A'}</div>
        </div>
        """

    @staticmethod
    def _generate_divider():
        # Matching your 'outline-Color-Neutral-200'
        return '<div style="height: 1px; background-color: #e2e8f0; width: 100%;"></div>'
